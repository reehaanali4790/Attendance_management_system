import datetime
import io
import re
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import DailyAttendance, DeviceSettings, Employee
from sync_service import is_working_day


THIN = Side(style="thin")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADERS = ("Date", "Name", "IN", "OUT", "Late", "Absent")


def _fmt_date(day: datetime.date) -> str:
    return day.strftime("%d/%m/%Y")


def _fmt_time(value: Optional[datetime.datetime]) -> str:
    if not value:
        return ""
    clock = value.time()
    return f"{clock.hour}:{clock.minute:02d}:{clock.second:02d}"


def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "", name).strip() or "Employee"
    return cleaned[:31]


def _apply_row_style(ws, row_idx: int, highlight: bool = False) -> None:
    for col in range(1, 7):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = CELL_BORDER
        cell.alignment = CENTER
        if highlight:
            cell.fill = BLUE_FILL


def _is_company_holiday(day: datetime.date, holiday_dates: set[datetime.date]) -> bool:
    return day in holiday_dates


def _is_late(record: Optional[DailyAttendance]) -> bool:
    return bool(record and record.status == "Late")


def _is_absent(
    record: Optional[DailyAttendance],
    day: datetime.date,
    today: datetime.date,
    holiday_dates: set[datetime.date],
) -> bool:
    if day in holiday_dates:
        return False
    if record and record.status in {"Absent", "On Leave"}:
        return True
    if day >= today:
        return False
    if record:
        return False
    return True


def build_individual_attendance_workbook(
    employee: Employee,
    records_by_date: dict[datetime.date, DailyAttendance],
    settings: DeviceSettings,
    start_date: datetime.date,
    end_date: datetime.date,
    holiday_dates: Optional[set[datetime.date]] = None,
) -> io.BytesIO:
    holiday_dates = holiday_dates or set()
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(employee.name)

    for col_idx, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = CENTER

    today = datetime.date.today()
    row_idx = 2
    total_late = 0
    total_absent = 0

    day = start_date
    while day <= end_date:
        is_holiday = _is_company_holiday(day, holiday_dates)
        if not is_holiday and not is_working_day(day, settings):
            day += datetime.timedelta(days=1)
            continue

        record = records_by_date.get(day)
        highlight = day.weekday() == 5 or is_holiday
        late_mark = ""
        absent_mark = ""

        if is_holiday:
            ws.cell(row=row_idx, column=1, value=_fmt_date(day))
            ws.cell(row=row_idx, column=2, value=employee.name)
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4)
            holiday_cell = ws.cell(row=row_idx, column=3, value="HOLIDAY")
            holiday_cell.alignment = CENTER
        else:
            late_flag = _is_late(record)
            absent_flag = _is_absent(record, day, today, holiday_dates)
            if late_flag:
                late_mark = "YES"
                total_late += 1
            if absent_flag:
                absent_mark = "YES"
                total_absent += 1

            ws.cell(row=row_idx, column=1, value=_fmt_date(day))
            ws.cell(row=row_idx, column=2, value=employee.name)
            ws.cell(row=row_idx, column=3, value=_fmt_time(record.check_in) if record else "")
            ws.cell(row=row_idx, column=4, value=_fmt_time(record.check_out) if record else "")
            ws.cell(row=row_idx, column=5, value=late_mark)
            ws.cell(row=row_idx, column=6, value=absent_mark)

        _apply_row_style(ws, row_idx, highlight=highlight)
        row_idx += 1
        day += datetime.timedelta(days=1)

    summary_row = row_idx + 1
    ws.cell(row=summary_row, column=5, value="TOTAL LATE").alignment = Alignment(horizontal="right")
    ws.cell(row=summary_row, column=5).font = HEADER_FONT
    late_total_cell = ws.cell(row=summary_row, column=6, value=total_late)
    late_total_cell.alignment = CENTER
    late_total_cell.border = CELL_BORDER
    late_total_cell.font = HEADER_FONT

    ws.cell(row=summary_row + 1, column=5, value="ABSENTS").alignment = Alignment(horizontal="right")
    ws.cell(row=summary_row + 1, column=5).font = HEADER_FONT
    absent_total_cell = ws.cell(row=summary_row + 1, column=6, value=total_absent)
    absent_total_cell.alignment = CENTER
    absent_total_cell.border = CELL_BORDER
    absent_total_cell.font = HEADER_FONT

    widths = (14, 18, 12, 12, 8, 8)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_bulk_attendance_workbook(records: list[DailyAttendance]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"
    headers = [
        "Date",
        "Employee Name",
        "Department",
        "Device ID",
        "Check In",
        "Check Out",
        "Work Hours",
        "Status",
        "Late Minutes",
        "Early Leave Minutes",
        "Remarks",
    ]
    ws.append(headers)

    for record in records:
        ws.append(
            [
                record.date.strftime("%Y-%m-%d"),
                record.employee.name,
                record.employee.department.name if record.employee.department else "",
                record.employee.user_id,
                record.check_in.strftime("%Y-%m-%d %H:%M:%S") if record.check_in else "",
                record.check_out.strftime("%Y-%m-%d %H:%M:%S") if record.check_out else "",
                record.work_hours,
                record.status,
                record.late_minutes,
                record.early_leave_minutes,
                record.remarks or "",
            ]
        )

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 10)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def individual_export_filename(employee: Employee, start_date: datetime.date, end_date: datetime.date) -> str:
    safe_name = re.sub(r"[^\w\-]+", "_", employee.name.strip()) or "employee"
    if start_date.year == end_date.year and start_date.month == end_date.month:
        month_label = start_date.strftime("%B_%Y")
        return f"{safe_name}_attendance_{month_label}.xlsx"
    return f"{safe_name}_attendance_{start_date}_to_{end_date}.xlsx"
